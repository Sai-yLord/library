from flask import Flask, render_template, request
from openpyxl import load_workbook
from sqlalchemy.orm.session import sessionmaker
from sqlalchemy.orm import scoped_session
from database import engine, Book

app = Flask(__name__)
db = scoped_session(sessionmaker(bind=engine))

@app.route("/")
def homepage():
    
    return render_template("index.html")

@app.route("/form/")
def form():
    return render_template("form.html")


@app.route("/books/")
def books():
    if 'key_word' in request.args:
        key_word = request.args.get("key_word")
        session = sessionmaker(engine)()
        books = session.execute(f"""
            SELECT *
            FROM "Book"
            WHERE name LIKE '%{key_word}%'
                    OR author ILIKE '%{key_word}%';
            """)
        session.commit()
    else:
        books = db.execute("""SELECT * FROM "Book";""")
        db.commit()

    return render_template("books_table.html", object_list=books)
#Что то с чем то ;)
# @app.route("/authors/")
# def authors():
#     excel = load_workbook("tales1.xlsx")
#     page = excel["Лист1"]
#     authors = {author.value for author in page["B"][1:]}
#     return render_template("authors.html", authors=authors)

@app.route("/authors/")
def authors():
    with engine.connect() as con:
        authors = con.execute('SELECT DISTINCT author FROM "Book" ;') 
    return render_template(
        "database_authors.html", authors=authors
    )
    

@app.route("/add/", methods=["POST"])
def add():
    f = request.form
    book = f["book"]
    author = f["author"]
    url = f["url"]    
    
    ids = db.execute('SELECT id FROM "Book" ORDER BY id DESC;')
    max_id = ids.first().id
    c_id = max_id + 1

    db.execute(f'''
        INSERT INTO "Book" (id, name, author, image)
        VALUES ({c_id}, '{book}', '{author}', '{url}');''')
    db.commit()

    message = "Form Received!"
    
    return render_template("index.html", message = message)

# @app.route("/db/add/", methods=["POST"])
# def db_add():
#     f = request.form
#     book = f["book"]
#     author = f["author"]
#     url = f["url"]

#     ids= db.execute('SELECT id FROM "Book" ORDER BY id DESC;')
#     max_id = ids.first().id
#     c_id = max_id + 1

#     with engine.connect() as con:
#         add = con.execute(f"""
#         INSERT INTO "Book" (id, name, author, image) 
#         VALUES ({c_id}, '{book}', '{author}', {url});""")
#     return "Form Accepted!"


@app.route("/book/<num>/")
def book(num):
    excel = load_workbook("tales1.xlsx")
    page = excel["Лист1"]
    object_list = [[tale.value, tale.offset(column=1).value, tale.offset(column=2).value] for tale in page["A"][1:]]
    obj = object_list[int(num)]
    obj.append(num)
    return render_template("book.htm", obj=obj)


@app.route("/db/book/<id>/")
def db_book(id):
    id = int(id)
    obj = db.execute(f'SELECT * FROM "Book" WHERE id = {id};').first()
    db.commit()
    return render_template("database_book.html", obj=obj)

@app.route("/book/<num>/edit/")
def book_edit(num):
    num = int(num) + 2                                                                        
    excel_file = load_workbook("tales1.xlsx")
    page = excel_file["Лист1"]
    tale = page[f"A{num}"]
    author = page[f"B{num}"]
    image = page[f"C{num}"]
    obj = [tale.value, author.value, image.value, num]
    return render_template("book_edit.html", obj=obj)

# @app.route("/book/<num>/save/", methods=["POST"])
# def book_save(num):
#     num = int(num)
#     excel_file = load_workbook("tales1.xlsx")
#     page = excel_file["Лист1"]
#     form = request.form
#     page[f"A{num}"] = form["tale"]
#     page[f"B{num}"] = form["author"]
#     page[f"C{num}"] = form["image"]
#     excel_file.save("tales1.xlsx")
#     return "Сохранено!"

@app.route("/<int:id>/", methods=["GET", "POST"])
def db_book_update(id):
    message = ''
    if request.method == "POST":
        name = request.form.get("tale")
        author = request.form.get("author")
        image = request.form.get("image")
        db.execute(f'''
            UPDATE "Book"
            SET 
                name='{name}',
                author='{author}',
                image='{image}'
            WHERE id={id};
        ''')
        db.commit()
        message = "Изменения сохранены"

    book_object = db.execute(f'SELECT * FROM "Book" WHERE id={id};').first()
    return render_template("database_book_update.html", book_object=book_object, message=message)